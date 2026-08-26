"""
Collecteur d'abonnés mobile PAR OPÉRATEUR — ANRT (Maroc), portail Open Data.

D'OÙ VIENT LA DONNÉE, ET POURQUOI UN XLSX PLUTÔT QU'UNE PAGE
    L'ANRT publie sur `data.gov.ma` un fichier XLSX unique, mis à jour à
    chaque trimestre : « Parc de la téléphonie mobile ». Vérifié le 23 août
    2026 en le téléchargeant : il couvre T1-2006 à T2-2026 (donc le trimestre
    QUI VIENT DE SE CLORE, pas un retard d'un an comme la Banque Mondiale /
    UIT), avec un parc PAR OPÉRATEUR
    (ITISSALAT AL-MAGHRIB, MEDI TELECOM, WANA CORPORATE) en plus du total.

CE QUE LE FICHIER CONTIENT VRAIMENT — VÉRIFIÉ, PAS SUPPOSÉ
    La feuille unique porte SEIZE colonnes légendées A→N sur les 14 premières
    lignes (une ligne = une colonne, ex. « I : Parc téléphonie mobile_ITISSALAT
    AL-MAGHRIB (en milliers) »), puis un tableau à partir de la ligne 16
    (l'en-tête réutilise les mêmes lettres) jusqu'à la dernière période. On
    lit la légende PLUTÔT QUE DE FIGER DES INDICES DE COLONNE : si l'ANRT
    réordonne ses colonnes d'un trimestre à l'autre, la légende continue de
    désigner le bon opérateur ; un indice figé se tromperait sans avertir.

    Les colonnes retenues sont UNIQUEMENT celles dont le libellé commence par
    « Parc téléphonie mobile_ » (le total prépayé+postpayé par opérateur) —
    PAS les colonnes "postpayé_" ou "prépayé_" séparées (D et H sont des
    sous-totaux globaux, pas par opérateur) ni la colonne globale (L).

    LES VALEURS SONT EN MILLIERS ("en milliers" dans le libellé même) :
    T2-2026 rend 18430 pour ITISSALAT AL-MAGHRIB, soit 18 430 000 abonnés — le
    même ordre de grandeur que Maroc Telecom dans la presse. Le collecteur
    multiplie par 1000 pour rendre un nombre d'abonnés, comme NCC Nigeria.

    Cellule "NA" = opérateur pas encore actif sur cette période (WANA
    CORPORATE n'existait pas en 2006) — écartée, pas convertie en zéro.

CE QUE « MEDI TELECOM » DÉSIGNE
    C'est le nom sous lequel l'ANRT continue de publier ce qui s'appelle
    commercialement Orange Maroc depuis le rachat par Orange S.A. (marque
    historique « Méditel »). `dim_subsidiary` ne connaît que « Orange Maroc »
    (code opérateur `orange`, pays `MA`, voir migration 003) — c'est donc là
    que ces lignes doivent atterrir, pas sous un opérateur « Medi Telecom »
    qui n'existe pas dans le modèle dimensionnel.

CE QUE CE MODULE N'EST PAS
    Comme `ncc_nigeria.py`, ce n'est pas un `BaseCollector` et il ne connaît
    pas le modèle dimensionnel : il rend un nom d'opérateur PUBLIÉ PAR LA
    SOURCE, charge à `OperatorMarketRepository` de le résoudre.
"""

import logging
import re
from io import BytesIO
from typing import Optional

import openpyxl
import requests

logger = logging.getLogger(__name__)

#: Fichier direct, vérifié le 23 août 2026. Change de nom de ressource si
#: l'ANRT republie le jeu de données ; la page ci-dessous est le premier
#: endroit à recontrôler si ce lien se met à échouer.
_URL_XLSX = (
    "https://data.gov.ma/data/fr/dataset/94969b16-8818-4ec9-b9a2-1d0f4ba2258c/"
    "resource/48cef8fd-edf0-4ec7-94d6-cf0fb30f475f/download/parc_telephonie_mobile.xlsx"
)
_URL_PAGE = (
    "https://data.gov.ma/data/fr/dataset/parc-de-la-telephonie-mobile-2006-2022/"
    "resource/48cef8fd-edf0-4ec7-94d6-cf0fb30f475f"
)

#: Nom publié par l'ANRT dans la légende -> code `dim_operator`. Fermée
#: comme `OPERATEURS` dans `ncc_nigeria.py` : un nom absent d'ici (une future
#: colonne ajoutée par l'ANRT) est ignoré plutôt que de faire échouer la
#: collecte.
OPERATEURS: dict[str, str] = {
    "ITISSALAT AL-MAGHRIB": "maroc_telecom",
    "MEDI TELECOM": "orange",
    "WANA CORPORATE": "inwi",
}

#: Préfixe et suffixe qui identifient, dans la légende, une colonne « total
#: par opérateur » — et l'excluent des sous-totaux (postpayé/prépayé) et du
#: total national.
_PREFIXE_LEGENDE = "Parc téléphonie mobile_"
_SUFFIXE_LEGENDE = " (en milliers)"

#: Trimestre -> (mois, dernier jour du mois). La source elle-même mesure
#: « à fin » de trimestre (en-tête « Période (à fin) », vérifié dans le
#: classeur) — voir `_periode`.
_FIN_TRIMESTRE = {1: (3, 31), 2: (6, 30), 3: (9, 30), 4: (12, 31)}

METRIC = "abonnes_mobile"
FREQUENCY = "quarterly"
SOURCE = "anrt_maroc"
ISO2 = "MA"

TIMEOUT = 30


class AnrtMarocCollector:
    """Interroge le jeu de données XLSX « Parc de la téléphonie mobile » de l'ANRT."""

    def __init__(self, timeout: int = TIMEOUT):
        self.timeout = timeout
        self.logger = logger

    def collect(self) -> tuple[list[dict], list[str]]:
        """Récupère le parc mobile trimestriel des 3 opérateurs marocains.

        Returns:
            (lignes prêtes pour `OperatorMarketRepository.upsert`, erreurs).
        """
        try:
            classeur = self._fetch()
        except Exception as e:  # noqa: BLE001
            return [], [f"ANRT Maroc : {e}"]

        try:
            feuille = classeur[classeur.sheetnames[0]]
        except Exception as e:  # noqa: BLE001
            return [], [f"ANRT Maroc : feuille introuvable — {e}"]

        legende = self._lire_legende(feuille)
        colonnes = self._colonnes_operateurs(legende)
        if not colonnes:
            return [], [
                "ANRT Maroc : aucune colonne « Parc téléphonie mobile_<opérateur> » "
                "trouvée — la légende a peut-être changé de forme"
            ]

        ligne_entete = self._trouver_ligne_entete(feuille)
        if ligne_entete is None:
            return [], ["ANRT Maroc : ligne d'en-tête (« Période (à fin) ») introuvable"]

        lettre_vers_index = self._index_colonnes(feuille, ligne_entete)
        # Résolu UNE FOIS avant la boucle sur les lignes : sinon chaque
        # période reparcourrait la légende pour chacun des 3 opérateurs.
        index_par_operateur = {
            code_operateur: lettre_vers_index.get(colonnes_lettre(legende, nom_source))
            for nom_source, code_operateur in colonnes.items()
        }

        lignes: list[dict] = []
        erreurs: list[str] = []
        for row in feuille.iter_rows(min_row=ligne_entete + 1, values_only=True):
            periode = self._periode(row[0] if row else None)
            if periode is None:
                continue  # ligne vide en fin de feuille, ou libellé illisible
            for code_operateur, index in index_par_operateur.items():
                if index is None or index >= len(row):
                    continue
                valeur = self._nombre(row[index])
                if valeur is None:
                    continue  # "NA" : opérateur pas encore actif sur ce trimestre
                lignes.append({
                    "operator_code": code_operateur,
                    "iso2": ISO2,
                    "metric": METRIC,
                    "period": periode,
                    "frequency": FREQUENCY,
                    "value": valeur * 1000,  # la source exprime "en milliers"
                    "source": SOURCE,
                    "source_url": _URL_PAGE,
                })

        self.logger.info(
            "ANRT Maroc : %d mesure(s) sur %d opérateur(s), %d erreur(s)",
            len(lignes), len(colonnes), len(erreurs),
        )
        return lignes, erreurs

    # ------------------------------------------------------------- Interne

    def _fetch(self):
        reponse = requests.get(
            _URL_XLSX,
            headers={"User-Agent": "Mozilla/5.0 (compatible; dashboard-veille/1.0)"},
            timeout=self.timeout,
        )
        reponse.raise_for_status()
        return openpyxl.load_workbook(BytesIO(reponse.content), data_only=True)

    @staticmethod
    def _lire_legende(feuille) -> dict[str, str]:
        """Lettre de colonne -> libellé complet, lu dans les premières lignes.

        S'ARRÊTE À LA PREMIÈRE LIGNE VIDE : la légende précède toujours un
        blanc puis la ligne d'en-tête (« Période (à fin) »). Continuer au-delà
        lirait des données comme si c'était de la légende.
        """
        legende: dict[str, str] = {}
        for row in feuille.iter_rows(min_row=1, max_col=2, values_only=True):
            lettre, libelle = (row + (None, None))[:2]
            if lettre is None or libelle is None:
                break
            if not isinstance(lettre, str) or len(lettre) > 2:
                break
            legende[lettre] = str(libelle)
        return legende

    @staticmethod
    def _colonnes_operateurs(legende: dict[str, str]) -> dict[str, str]:
        """Nom d'opérateur publié -> code `dim_operator`, pour les colonnes
        de TOTAL par opérateur seulement (voir la documentation du module)."""
        trouves: dict[str, str] = {}
        for libelle in legende.values():
            if not (libelle.startswith(_PREFIXE_LEGENDE) and libelle.endswith(_SUFFIXE_LEGENDE)):
                continue
            nom = libelle[len(_PREFIXE_LEGENDE):-len(_SUFFIXE_LEGENDE)]
            if nom in OPERATEURS:
                trouves[nom] = OPERATEURS[nom]
        return trouves

    @staticmethod
    def _trouver_ligne_entete(feuille) -> Optional[int]:
        for i, row in enumerate(feuille.iter_rows(min_row=1, max_col=1, values_only=True), 1):
            valeur = row[0] if row else None
            if isinstance(valeur, str) and valeur.strip().lower().startswith("période"):
                return i
        return None

    @staticmethod
    def _index_colonnes(feuille, ligne_entete: int) -> dict[str, int]:
        """Lettre -> index dans la ligne (0-based), lu sur LA LIGNE D'EN-TÊTE
        elle-même — pas la position dans le classeur, qui peut différer si
        une colonne a été insérée avant."""
        entete = next(feuille.iter_rows(min_row=ligne_entete, max_row=ligne_entete, values_only=True))
        return {v: i for i, v in enumerate(entete) if isinstance(v, str)}

    @staticmethod
    def _periode(brut) -> Optional[str]:
        """« T2-2026 » -> « 2026-06-30 » (DERNIER jour du trimestre).

        PAS le premier jour. La source mesure un parc « à fin » de trimestre
        — c'est écrit dans son propre en-tête, « Période (à fin) ». Stocker
        le premier jour afficherait « avril 2026 » pour une valeur measurée
        fin juin, soit deux mois d'écart avec la réalité — dans le sens qui
        fait paraître la donnée PLUS ancienne qu'elle ne l'est.
        """
        if not isinstance(brut, str):
            return None
        m = re.match(r"T([1-4])-(\d{4})", brut.strip())
        if not m:
            return None
        trimestre, annee = int(m.group(1)), m.group(2)
        mois, dernier_jour = _FIN_TRIMESTRE[trimestre]
        return f"{annee}-{mois:02d}-{dernier_jour:02d}"

    @staticmethod
    def _nombre(brut) -> Optional[float]:
        if isinstance(brut, (int, float)):
            return float(brut)
        return None


def colonnes_lettre(legende: dict[str, str], nom_operateur: str) -> Optional[str]:
    """Lettre de colonne portant `nom_operateur`, ou None. Fonction à part
    (plutôt qu'une fermeture) pour rester testable isolément."""
    cherche = f"{_PREFIXE_LEGENDE}{nom_operateur}{_SUFFIXE_LEGENDE}"
    for lettre, libelle in legende.items():
        if libelle == cherche:
            return lettre
    return None
